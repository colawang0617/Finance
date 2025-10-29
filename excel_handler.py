"""
Excel Handler Module

Manages all Excel file operations including:
- Finding next empty row
- Inserting daily data
- Preserving cell styling (colors, fonts, alignment)
- Generating formulas for calculated columns
- Creating backups
"""

import os
import shutil
from datetime import datetime
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelHandlerError(Exception):
    """Raised when Excel operations fail"""
    pass


class ExcelHandler:
    """Handle Excel file operations with styling preservation"""

    # Column mapping: Excel column -> data key
    COLUMN_MAP = {
        'A': 'date',
        'B': 'FORMULA',  # =C+D+E+F+G+H
        'C': 'meituan',
        'D': 'stored_card_redemption',
        'E': 'douyin',
        'F': 'coaching_redemption',
        'G': 'wechat',
        'H': 'alipay',
        'I': 'FORMULA',  # =J+K+L
        'J': 'water',
        'K': 'gatorade',
        'L': 'other',
        'M': 'trial_class',
        'N': 'stored_card_recharge',
        'O': 'private_coaching_recharge',
        'P': 'monthly_card',
        'Q': 'FORMULA',  # =B+I+P
        'R': 'FORMULA'   # =B+I+M+N+O+P
    }

    # Styling configuration for each column group
    STYLES = {
        'A': {  # Date
            'fill': PatternFill(patternType='solid', fgColor='FFB4A7D6'),
            'font': Font(name='Cambria', size=11, color='FFFFFFFF'),
            'alignment': Alignment(horizontal='center')
        },
        'B': {  # Venue total (formula)
            'fill': PatternFill(patternType='solid', fgColor='FF366092'),
            'font': Font(name='Cambria', size=11, color='FFFFFF00'),
            'alignment': Alignment(horizontal='center')
        },
        'C-H': {  # Venue subcategories
            'fill': PatternFill(patternType='solid', fgColor='FF5B9BD5'),
            'font': Font(name='Cambria', size=11, color='FF000000'),
            'alignment': Alignment(horizontal='center')
        },
        'I': {  # Store total (formula)
            'fill': PatternFill(patternType='solid', fgColor='FF366092'),
            'font': Font(name='Cambria', size=11, color='FFFFFF00'),
            'alignment': Alignment(horizontal='center')
        },
        'J-L': {  # Store items
            'fill': PatternFill(patternType='solid', fgColor='FF5B9BD5'),
            'font': Font(name='Cambria', size=11, color='FF000000'),
            'alignment': Alignment(horizontal='center')
        },
        'M-P': {  # Recharges
            'fill': PatternFill(patternType='solid', fgColor='FF366092'),
            'font': Font(name='Cambria', size=11, color='FFFFFF00'),
            'alignment': Alignment(horizontal='center')
        },
        'Q-R': {  # Totals
            'fill': PatternFill(patternType='solid', fgColor='FFF4CCCC'),
            'font': Font(name='Cambria', size=11, color='FFCC0000'),
            'alignment': Alignment(horizontal='center')
        }
    }

    def __init__(self, filepath: str):
        """
        Initialize Excel handler

        Args:
            filepath: Path to Excel file

        Raises:
            ExcelHandlerError: If file doesn't exist or cannot be opened
        """
        if not os.path.exists(filepath):
            raise ExcelHandlerError(f"Excel文件不存在: {filepath}")

        self.filepath = filepath
        self.workbook = None
        self.sheet = None

        try:
            self.workbook = openpyxl.load_workbook(filepath)
            if '每日数据' not in self.workbook.sheetnames:
                raise ExcelHandlerError("Excel文件中未找到'每日数据'工作表")
            self.sheet = self.workbook['每日数据']
        except PermissionError:
            raise ExcelHandlerError("无法访问Excel文件，请确认文件未被其他程序占用")
        except Exception as e:
            raise ExcelHandlerError(f"打开Excel文件失败: {str(e)}")

    def find_next_row(self) -> int:
        """
        Find the next empty row in the sheet

        Returns:
            Row number for insertion (1-indexed)
        """
        # Start from row 3 (first data row after header and empty row)
        for row in range(3, self.sheet.max_row + 2):
            if self.sheet[f'A{row}'].value is None:
                return row
        return self.sheet.max_row + 1

    def check_duplicate_date(self, date: str) -> Optional[int]:
        """
        Check if date already exists in the sheet

        Args:
            date: Date in MM-DD format

        Returns:
            Row number if date exists, None otherwise
        """
        for row in range(3, self.sheet.max_row + 1):
            cell_value = self.sheet[f'A{row}'].value
            if cell_value == date:
                return row
        return None

    def generate_formulas(self, row: int) -> Dict[str, str]:
        """
        Generate Excel formulas for calculated columns

        Args:
            row: Row number (1-indexed)

        Returns:
            Dictionary mapping column letters to formula strings
        """
        return {
            'B': f'=C{row}+D{row}+E{row}+F{row}+G{row}+H{row}',
            'I': f'=J{row}+K{row}+L{row}',
            'Q': f'=B{row}+I{row}+P{row}',
            'R': f'=B{row}+I{row}+M{row}+N{row}+O{row}+P{row}'
        }

    def _get_style_for_column(self, col: str) -> Dict[str, Any]:
        """Get the appropriate style for a column"""
        # Direct match
        if col in self.STYLES:
            return self.STYLES[col]

        # Range match (e.g., C-H)
        for style_key, style_value in self.STYLES.items():
            if '-' in style_key:
                start, end = style_key.split('-')
                if start <= col <= end:
                    return style_value

        # Default style (shouldn't reach here)
        return self.STYLES['C-H']  # Default to basic style

    def apply_cell_styling(self, cell, col_letter: str):
        """
        Apply styling to a cell based on its column

        Args:
            cell: openpyxl Cell object
            col_letter: Column letter (A, B, C, etc.)
        """
        style = self._get_style_for_column(col_letter)

        # Create NEW style objects (don't reuse) to avoid openpyxl issues
        cell.fill = PatternFill(
            patternType=style['fill'].patternType,
            fgColor=style['fill'].fgColor
        )
        cell.font = Font(
            name=style['font'].name,
            size=style['font'].size,
            color=style['font'].color,
            bold=style['font'].bold
        )
        cell.alignment = Alignment(
            horizontal=style['alignment'].horizontal
        )

    def insert_daily_data(self, data: Dict[str, Any]) -> int:
        """
        Insert parsed daily data into the next available row

        Args:
            data: Parsed data dictionary from parser

        Returns:
            Row number where data was inserted

        Raises:
            ExcelHandlerError: If insertion fails
        """
        try:
            # Find next empty row
            row = self.find_next_row()

            # Generate formulas
            formulas = self.generate_formulas(row)

            # Insert data for each column
            for col_letter, data_key in self.COLUMN_MAP.items():
                cell = self.sheet[f'{col_letter}{row}']

                if data_key == 'FORMULA':
                    # Insert formula
                    cell.value = formulas[col_letter]
                else:
                    # Insert data value (or None)
                    cell.value = data.get(data_key)

                # Apply styling
                self.apply_cell_styling(cell, col_letter)

            return row

        except Exception as e:
            raise ExcelHandlerError(f"插入数据失败: {str(e)}")

    def create_backup(self) -> str:
        """
        Create a timestamped backup of the Excel file

        Returns:
            Path to backup file

        Raises:
            ExcelHandlerError: If backup creation fails
        """
        try:
            # Generate backup filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_name = os.path.basename(self.filepath)
            name_without_ext = os.path.splitext(base_name)[0]
            ext = os.path.splitext(base_name)[1]
            backup_name = f"{name_without_ext}_backup_{timestamp}{ext}"
            backup_path = os.path.join(os.path.dirname(self.filepath), backup_name)

            # Copy file
            shutil.copy2(self.filepath, backup_path)
            return backup_path

        except Exception as e:
            raise ExcelHandlerError(f"创建备份失败: {str(e)}")

    def save(self):
        """
        Save changes to the Excel file

        Raises:
            ExcelHandlerError: If save fails
        """
        try:
            self.workbook.save(self.filepath)
        except PermissionError:
            raise ExcelHandlerError("无法保存文件，请确认文件未被其他程序占用")
        except Exception as e:
            raise ExcelHandlerError(f"保存文件失败: {str(e)}")

    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()

    def __enter__(self):
        """Context manager entry"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        self.close()


if __name__ == '__main__':
    # Test with sample data
    test_data = {
        'date': '10-28',
        'meituan': 144,
        'stored_card_redemption': 505,
        'douyin': None,
        'coaching_redemption': 90,
        'wechat': None,
        'alipay': None,
        'water': None,
        'gatorade': None,
        'other': None,
        'trial_class': None,
        'stored_card_recharge': 1000,
        'private_coaching_recharge': None,
        'monthly_card': None
    }

    try:
        handler = ExcelHandler('Finance/财务跟踪表_完整版_KL.xlsx')
        print(f"下一个空行: {handler.find_next_row()}")

        # Check for duplicate
        dup_row = handler.check_duplicate_date(test_data['date'])
        if dup_row:
            print(f"警告: 日期 {test_data['date']} 已存在于第 {dup_row} 行")
        else:
            print(f"日期 {test_data['date']} 未重复")

        print("\n测试完成 (未实际插入数据)")
        handler.close()

    except ExcelHandlerError as e:
        print(f"错误: {e}")
