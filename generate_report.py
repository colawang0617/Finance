#!/usr/bin/env python3
"""
Financial Report Generator

Easy-to-use script for generating financial visualization reports
"""

import sys
import os
import openpyxl
from visualizations import monthly_category, daily_progression, monthly_pies, statistical_analysis


def get_available_months(filepath=None):
    """
    Automatically detect which months have data in the Excel file

    Args:
        filepath: Path to Excel file (defaults to standard location)

    Returns:
        List of month numbers that have data (e.g., [5, 6, 7, 8, 9, 10, 11])
    """
    if filepath is None:
        # Default path
        base_dir = os.path.dirname(os.path.abspath(__file__))
        filepath = os.path.join(base_dir, 'Finance', '财务跟踪表_完整版_KL.xlsx')

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb['每日数据']

        # Collect unique months from date column
        months = set()
        for row in range(3, ws.max_row + 1):
            date_val = ws[f'A{row}'].value
            if date_val and isinstance(date_val, str) and '-' in date_val:
                # Parse "MM-DD" format
                month = int(date_val.split('-')[0])
                months.add(month)

        wb.close()

        # Return sorted list
        return sorted(list(months))

    except Exception as e:
        print(f"Warning: Could not auto-detect months ({e}), using default range")
        # Fallback to extended default range
        return [5, 6, 7, 8, 9, 10, 11, 12]


def print_header():
    """Print header"""
    print("\n" + "=" * 60)
    print("           财务可视化报告生成器")
    print("        Financial Visualization Generator")
    print("=" * 60)


def print_menu():
    """Display menu"""
    print("\n请选择要生成的图表:")
    print("-" * 60)
    print("1. 月度分类进展图       (Monthly Category Progression)")
    print("2. 每日场地进展图       (Daily Venue Progression)")
    print("3. 月度百分比分布图     (Monthly Percentage Distribution)")
    print("4. 统计分析图           (Statistical Analysis) ⭐")
    print("5. 生成所有图表         (Generate All Charts)")
    print("0. 退出                 (Exit)")
    print("-" * 60)


def generate_all_charts():
    """Generate all visualization charts"""
    print("\n正在生成所有图表...")
    print("-" * 60)

    # Auto-detect available months
    available_months = get_available_months()
    print(f"检测到的月份: {', '.join(map(str, available_months))}")

    try:
        # Chart 1: Monthly Category Progression
        print("\n[1/4] 正在生成月度分类进展图...")
        file1 = monthly_category.generate_monthly_category_chart(
            start_month=min(available_months),
            end_month=max(available_months)
        )

        # Chart 2: Daily Venue Progression
        print("\n[2/4] 正在生成每日场地进展图...")
        file2 = daily_progression.generate_daily_progression(
            months=available_months
        )

        # Chart 3: Monthly Percentage Distribution
        print("\n[3/4] 正在生成月度百分比分布图...")
        file3 = monthly_pies.generate_monthly_pies(
            months=available_months
        )

        # Chart 4: Statistical Analysis
        print("\n[4/4] 正在生成统计分析图...")
        file4 = statistical_analysis.generate_statistical_analysis(
            months=available_months
        )

        print("\n" + "=" * 60)
        print("           ✓ 所有图表已成功生成!")
        print("=" * 60)
        print(f"\n生成的文件:")
        print(f"  1. {file1}")
        print(f"  2. {file2}")
        print(f"  3. {file3}")
        print(f"  4. {file4}")

        return True

    except Exception as e:
        print(f"\n✗ 生成失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main function"""
    print_header()

    # Auto-detect available months
    available_months = get_available_months()
    default_start = min(available_months)
    default_end = max(available_months)
    default_months_str = ','.join(map(str, available_months))

    while True:
        print_menu()
        choice = input("\n请选择 (0-4): ").strip()

        try:
            if choice == '1':
                print("\n正在生成月度分类进展图...")
                start = input(f"起始月份 (默认{default_start}): ").strip() or str(default_start)
                end = input(f"结束月份 (默认{default_end}): ").strip() or str(default_end)

                file_path = monthly_category.generate_monthly_category_chart(
                    start_month=int(start),
                    end_month=int(end)
                )
                print(f"\n✓ 图表已生成: {file_path}")

            elif choice == '2':
                print("\n正在生成每日场地进展图...")
                months_input = input(f"月份列表 (用逗号分隔, 默认{default_months_str}): ").strip()

                if months_input:
                    months = [int(m.strip()) for m in months_input.split(',')]
                else:
                    months = available_months

                file_path = daily_progression.generate_daily_progression(months=months)
                print(f"\n✓ 图表已生成: {file_path}")

            elif choice == '3':
                print("\n正在生成月度百分比分布图...")
                months_input = input(f"月份列表 (用逗号分隔, 默认{default_months_str}): ").strip()

                if months_input:
                    months = [int(m.strip()) for m in months_input.split(',')]
                else:
                    months = available_months

                file_path = monthly_pies.generate_monthly_pies(months=months)
                print(f"\n✓ 图表已生成: {file_path}")

            elif choice == '4':
                print("\n正在生成统计分析图...")
                months_input = input(f"月份列表 (用逗号分隔, 默认{default_months_str}): ").strip()

                if months_input:
                    months = [int(m.strip()) for m in months_input.split(',')]
                else:
                    months = available_months

                file_path = statistical_analysis.generate_statistical_analysis(months=months)
                print(f"\n✓ 图表已生成: {file_path}")

            elif choice == '5':
                generate_all_charts()

            elif choice == '0':
                print("\n👋 再见!")
                break

            else:
                print("\n✗ 无效选项,请重新选择")

            if choice != '0' and choice in ['1', '2', '3', '4', '5']:
                input("\n按 Enter 继续...")

        except ValueError:
            print("\n✗ 输入错误,请输入有效的数字")
            input("\n按 Enter 继续...")
        except KeyboardInterrupt:
            print("\n\n👋 程序已退出")
            break
        except Exception as e:
            print(f"\n✗ 错误: {e}")
            import traceback
            traceback.print_exc()
            input("\n按 Enter 继续...")


def command_line_mode():
    """Command line argument mode"""
    import argparse

    parser = argparse.ArgumentParser(description='生成财务可视化报告')
    parser.add_argument('--all', action='store_true', help='生成所有图表')
    parser.add_argument('--monthly', action='store_true', help='生成月度分类进展图')
    parser.add_argument('--daily', action='store_true', help='生成每日进展图')
    parser.add_argument('--pies', action='store_true', help='生成月度百分比分布图')
    parser.add_argument('--stats', action='store_true', help='生成统计分析图')
    parser.add_argument('--months', type=str, help='月份 (逗号分隔, 如: 5,6,7,8,9,10,11)')

    args = parser.parse_args()

    # Parse months - auto-detect if not specified
    if args.months:
        months = [int(m.strip()) for m in args.months.split(',')]
    else:
        months = get_available_months()
        print(f"自动检测到的月份: {', '.join(map(str, months))}")

    # Generate charts
    if args.all:
        generate_all_charts()
    elif args.monthly:
        monthly_category.generate_monthly_category_chart(
            start_month=min(months),
            end_month=max(months)
        )
    elif args.daily:
        daily_progression.generate_daily_progression(months=months)
    elif args.pies:
        monthly_pies.generate_monthly_pies(months=months)
    elif args.stats:
        statistical_analysis.generate_statistical_analysis(months=months)
    else:
        # Interactive mode
        main()


if __name__ == '__main__':
    if len(sys.argv) > 1:
        # Command line mode
        command_line_mode()
    else:
        # Interactive mode
        main()
