#!/usr/bin/env python3
"""
Update Monthly Summary Formulas

This script updates the hardcoded formulas in the 月度汇总 sheet
to use dynamic SUMIFS formulas that automatically include all data
for each month based on the date in the 每日数据 sheet.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime


def create_backup(filepath):
    """Create a timestamped backup of the Excel file"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = filepath.replace('.xlsx', f'_backup_{timestamp}.xlsx')
    shutil.copy2(filepath, backup_path)
    print(f"✓ 备份已创建: {backup_path}")
    return backup_path


def update_monthly_formulas(filepath, month_number, month_column):
    """
    Update formulas for a specific month to use SUMIFS

    Args:
        filepath: Path to Excel file
        month_number: Month number (e.g., "10" for October)
        month_column: Column letter where the month data is (e.g., "I" for October)
    """
    wb = openpyxl.load_workbook(filepath)
    ws_summary = wb['月度汇总']

    # Column mapping for 每日数据
    data_columns = {
        '大众美团': 'C',
        '储值卡核销': 'D',
        '抖音': 'E',
        '教练课核销': 'F',
        '微信': 'G',
        '支付宝': 'H',
        '水': 'J',
        '佳得乐': 'K',
        '其他': 'L',
        '体验课': 'M',
        '储值卡充值': 'N',
        '私教课': 'O',
        '月卡': 'P'
    }

    # Find the month's starting row in 月度汇总
    month_row = None
    for row in range(1, ws_summary.max_row + 1):
        # Check both column A and F for month headers
        cell_f = ws_summary[f'F{row}'].value
        if cell_f and isinstance(cell_f, str) and cell_f == f'{month_number}月':
            month_row = row
            break

    if not month_row:
        print(f"✗ 未找到 {month_number}月 的数据")
        return False

    print(f"\n更新 {month_number}月 的公式 (行 {month_row}, 列 {month_column}):")
    print("-" * 80)

    # Update formulas for each category
    updates_made = 0

    for idx, (category, data_col) in enumerate(data_columns.items(), start=1):
        # Calculate the row for this category
        if category in ['大众美团', '储值卡核销', '抖音', '教练课核销', '微信', '支付宝']:
            # Venue subcategories start at row+2
            cat_row = month_row + idx
        elif category in ['水', '佳得乐', '其他']:
            # Store items - adjust offset
            offset = list(data_columns.keys()).index(category) + 1
            cat_row = month_row + offset + 1
        else:
            # Other categories
            offset = list(data_columns.keys()).index(category) + 1
            cat_row = month_row + offset + 1

        # Create new SUMIFS formula
        # Pattern: =SUMIFS(每日数据!C:C, 每日数据!A:A, "10-*")
        new_formula = f'=SUMIFS(每日数据!{data_col}:{data_col}, 每日数据!$A:$A, "{month_number}-*")'

        # Get current cell
        cell = ws_summary[f'{month_column}{cat_row}']
        old_formula = cell.value

        # Update only if it's a formula that references 每日数据
        if old_formula and isinstance(old_formula, str) and '每日数据' in old_formula:
            cell.value = new_formula
            print(f"  {category:12s} (行{cat_row}): 已更新")
            updates_made += 1

    if updates_made > 0:
        wb.save(filepath)
        print(f"\n✓ 成功更新 {updates_made} 个公式")
        return True
    else:
        print("\n⚠ 没有更新任何公式")
        return False

    wb.close()


def update_october_formulas_detailed(filepath):
    """
    Update October formulas with detailed row mapping
    """
    wb = openpyxl.load_workbook(filepath)
    ws_summary = wb['月度汇总']

    print("\n更新10月公式 (详细模式):")
    print("=" * 80)

    # October is at row 20, column I
    october_mappings = [
        (21, 'I', 'C', '大众美团'),
        (22, 'I', 'D', '储值卡核销'),
        (23, 'I', 'E', '抖音'),
        (24, 'I', 'F', '教练课核销'),
        (25, 'I', 'G', '微信'),
        (26, 'I', 'H', '支付宝'),
        (28, 'I', 'J', '水'),
        (29, 'I', 'K', '佳得乐'),
        (30, 'I', 'L', '其他'),
        (31, 'I', 'M', '体验课'),
        (32, 'I', 'N', '储值卡充值'),
        (33, 'I', 'O', '私教课'),
        (34, 'I', 'P', '月卡'),
    ]

    updates_made = 0
    for summary_row, summary_col, data_col, category in october_mappings:
        cell = ws_summary[f'{summary_col}{summary_row}']
        old_formula = cell.value

        # Create new SUMIFS formula
        new_formula = f'=SUMIFS(每日数据!{data_col}:{data_col},每日数据!$A:$A,"10-*")'

        # Update
        if old_formula and isinstance(old_formula, str):
            cell.value = new_formula
            print(f"  行{summary_row} {category:12s}: {old_formula[:40]} → SUMIFS")
            updates_made += 1

    wb.save(filepath)
    print(f"\n✓ 10月: 成功更新 {updates_made} 个公式")
    wb.close()


def main():
    """Main function"""
    filepath = 'Finance/财务跟踪表_完整版_KL.xlsx'

    print("=" * 80)
    print("        更新月度汇总公式工具")
    print("=" * 80)

    # Create backup
    create_backup(filepath)

    # Update October formulas (the most recent month with hardcoded ranges)
    update_october_formulas_detailed(filepath)

    print("\n" + "=" * 80)
    print("           ✓ 更新完成!")
    print("=" * 80)
    print("\n现在10月的公式会自动包含所有 10-XX 格式的日期数据。")
    print("当您添加新的10月数据时，月度汇总会自动更新。")


if __name__ == '__main__':
    main()
