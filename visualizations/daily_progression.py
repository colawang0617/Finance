"""
Daily Progression Chart for Venue Income

Generates a line chart showing daily venue income with clear data labels
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib
from datetime import datetime
import os
import numpy as np

# Set Chinese font support - Use Heiti TC for bold, rounded style
matplotlib.rcParams['font.sans-serif'] = ['Heiti TC', 'STHeiti', 'Microsoft YaHei', 'SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False
matplotlib.rcParams['font.size'] = 12
matplotlib.rcParams['font.weight'] = 'medium'  # Bolder default weight


def get_daily_venue_data(filepath, months):
    """
    Extract daily venue income for specified months

    Args:
        filepath: Path to Excel file
        months: List of months to include (e.g., [5, 6, 7, 8, 9, 10])

    Returns:
        Dictionary with dates and values by month
    """
    # Use data_only=True to get calculated values instead of formulas
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['每日数据']

    # Initialize data structure
    daily_data = {month: {'dates': [], 'values': []} for month in months}

    # Collect data
    for row in range(3, ws.max_row + 1):
        date_val = ws[f'A{row}'].value
        if not date_val:
            continue

        try:
            # Handle both string dates ("08-18") and datetime objects
            if isinstance(date_val, str):
                month = int(date_val.split('-')[0])
                date_str = date_val
            else:
                # It's a datetime object
                month = date_val.month
                date_str = f'{date_val.month:02d}-{date_val.day:02d}'

            if month not in months:
                continue

            venue = ws[f'B{row}'].value or 0
            if isinstance(venue, (int, float)):
                daily_data[month]['dates'].append(date_str)
                daily_data[month]['values'].append(venue)

        except (ValueError, IndexError):
            continue

    wb.close()
    return daily_data


def generate_daily_progression(months=[5, 6, 7, 8, 9, 10], filepath='Finance/财务跟踪表_完整版_KL.xlsx', output_dir='reports/graphs'):
    """
    Generate daily venue income progression chart with trend lines and data labels

    Args:
        months: List of months to display
        filepath: Path to Excel file
        output_dir: Directory to save output PNG

    Returns:
        Path to generated PNG file
    """
    # Get data
    data = get_daily_venue_data(filepath, months)

    # Filter out empty months
    data = {m: v for m, v in data.items() if len(v['values']) > 0}

    if not data:
        print("✗ 没有找到数据")
        return None

    # Month colors
    month_colors = {
        5: '#E57373',   # May - Red
        6: '#4DB6AC',   # June - Cyan
        7: '#4FC3F7',   # July - Light Blue
        8: '#FFB74D',   # August - Orange
        9: '#81C784',   # September - Green
        10: '#9575CD'   # October - Purple
    }

    month_names_zh = {
        5: '5月', 6: '6月', 7: '7月',
        8: '8月', 9: '9月', 10: '10月'
    }

    # Calculate overall mean
    all_values = []
    for month_data in data.values():
        all_values.extend(month_data['values'])
    overall_mean = np.mean(all_values) if all_values else 0

    # Create figure - clean and concise
    fig, ax = plt.subplots(figsize=(20, 8))
    fig.suptitle('场地入账 - 每日进展 (2025)', fontsize=20, fontweight='bold')

    # Plot data for each month
    x_position = 0
    all_x_positions = []
    all_labels = []

    for month in sorted(data.keys()):
        month_data = data[month]
        dates = month_data['dates']
        values = month_data['values']
        color = month_colors.get(month, '#666666')

        # X positions for this month
        x_positions = list(range(x_position, x_position + len(dates)))
        all_x_positions.extend(x_positions)

        # Plot line with markers - clean style
        ax.plot(x_positions, values, marker='o', color=color,
               linewidth=3, markersize=7, alpha=0.85,
               label=month_names_zh.get(month, f'{month}月'),
               markeredgecolor='white', markeredgewidth=1.5)

        # Add data labels - ONLY for peak values to avoid clutter
        max_val = max(values)
        min_val = min(values)
        for i, (x, y) in enumerate(zip(x_positions, values)):
            # Only show labels for max, min, or every 7th point
            if y == max_val or y == min_val or (i % 7 == 0 and i > 0):
                ax.text(x, y, f'¥{y:,.0f}',
                       ha='center', va='bottom' if y > overall_mean else 'top',
                       fontsize=9, color=color, fontweight='bold',
                       bbox=dict(boxstyle='round,pad=0.4', facecolor='white',
                                edgecolor=color, alpha=0.85, linewidth=1.5))

        # Store labels for x-axis
        all_labels.extend([d.split('-')[1] for d in dates])  # Just day number

        x_position += len(dates)

    # Draw subtle mean line (not too distracting)
    ax.axhline(y=overall_mean, color='#FF6B6B', linestyle=':', linewidth=2,
               label=f'平均值: ¥{overall_mean:,.2f}', alpha=0.6)

    # Formatting - All Chinese, clean
    ax.set_xlabel('日期', fontsize=13, fontweight='bold')
    ax.set_ylabel('金额 (元)', fontsize=13, fontweight='bold')
    ax.grid(axis='y', alpha=0.2, linestyle='--', linewidth=0.8)
    ax.grid(axis='x', alpha=0.1)
    ax.legend(loc='upper left', fontsize=11, framealpha=0.95, edgecolor='gray')

    # Set x-axis labels (show every 5th day)
    ax.set_xticks([i for i in range(len(all_x_positions)) if i % 5 == 0])
    ax.set_xticklabels([all_labels[i] for i in range(len(all_labels)) if i % 5 == 0], rotation=45, ha='right')

    # Format y-axis
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'¥{int(x):,}'))

    plt.tight_layout()

    # Save to file
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'每日进展_场地_{timestamp}.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"✓ 每日进展图已生成: {output_file}")
    return output_file


if __name__ == '__main__':
    # Test
    generate_daily_progression()
