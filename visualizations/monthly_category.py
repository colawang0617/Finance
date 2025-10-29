"""
Monthly Category Progression Chart

Generates a 6-panel bar chart showing monthly revenue progression
for different categories (Venue, Store, Trial Class, Card Topup, Private Coach, Monthly Card)
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib
from datetime import datetime
import os

# Set Chinese font support - Use Heiti TC for bold, rounded style
matplotlib.rcParams['font.sans-serif'] = ['Heiti TC', 'STHeiti', 'Microsoft YaHei', 'SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False
matplotlib.rcParams['font.size'] = 12
matplotlib.rcParams['font.weight'] = 'medium'  # Bolder default weight


def get_monthly_data(filepath, start_month, end_month):
    """
    Extract monthly totals for each category from Excel

    Args:
        filepath: Path to Excel file
        start_month: Starting month (1-12)
        end_month: Ending month (1-12)

    Returns:
        Dictionary with monthly totals for each category
    """
    # Use data_only=True to get calculated values instead of formulas
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['每日数据']

    # Initialize data structure
    monthly_data = {
        'venue': {},      # Column B
        'store': {},      # Column I
        'trial': {},      # Column M
        'card_topup': {}, # Column N
        'private_coach': {},  # Column O
        'monthly_card': {}    # Column P
    }

    # Collect data by month
    for row in range(3, ws.max_row + 1):
        date_val = ws[f'A{row}'].value
        if not date_val:
            continue

        # Handle both string dates ("08-18") and datetime objects
        try:
            if isinstance(date_val, str):
                month = int(date_val.split('-')[0])
            else:
                # It's a datetime object
                month = date_val.month

            if month < start_month or month > end_month:
                continue

            month_key = month

            # Get values for each category
            venue = ws[f'B{row}'].value or 0
            store = ws[f'I{row}'].value or 0
            trial = ws[f'M{row}'].value or 0
            card_topup = ws[f'N{row}'].value or 0
            private_coach = ws[f'O{row}'].value or 0
            monthly_card = ws[f'P{row}'].value or 0

            # Accumulate monthly totals
            if month_key not in monthly_data['venue']:
                for key in monthly_data:
                    monthly_data[key][month_key] = 0

            monthly_data['venue'][month_key] += venue if isinstance(venue, (int, float)) else 0
            monthly_data['store'][month_key] += store if isinstance(store, (int, float)) else 0
            monthly_data['trial'][month_key] += trial if isinstance(trial, (int, float)) else 0
            monthly_data['card_topup'][month_key] += card_topup if isinstance(card_topup, (int, float)) else 0
            monthly_data['private_coach'][month_key] += private_coach if isinstance(private_coach, (int, float)) else 0
            monthly_data['monthly_card'][month_key] += monthly_card if isinstance(monthly_card, (int, float)) else 0

        except (ValueError, IndexError):
            continue

    wb.close()
    return monthly_data


def generate_monthly_category_chart(start_month=5, end_month=10, filepath='Finance/财务跟踪表_完整版_KL.xlsx', output_dir='reports/graphs'):
    """
    Generate monthly progression chart for all categories

    Args:
        start_month: Starting month (default: 5 for May)
        end_month: Ending month (default: 10 for October)
        filepath: Path to Excel file
        output_dir: Directory to save output PNG

    Returns:
        Path to generated PNG file
    """
    # Get data
    data = get_monthly_data(filepath, start_month, end_month)

    # Prepare month labels
    months = list(range(start_month, end_month + 1))
    month_labels = [f'2025-{m:02d}' for m in months]

    # Create figure with 6 subplots (2 rows x 3 columns)
    fig, axes = plt.subplots(2, 3, figsize=(18, 10))
    fig.suptitle('各类别月度进展 (2025)', fontsize=22, fontweight='bold', y=0.995)

    # Category configurations - Chinese only
    categories = [
        ('venue', '场地入账金额', '#5DADE2'),
        ('store', '云店销售', '#C39BD3'),
        ('trial', '体验课', '#F4A460'),
        ('card_topup', '储值卡充值', '#CD6155'),
        ('private_coach', '私教课', '#82B366'),
        ('monthly_card', '月卡', '#E57373')
    ]

    # Plot each category
    for idx, (key, title, color) in enumerate(categories):
        row = idx // 3
        col = idx % 3
        ax = axes[row, col]

        # Get values for this category
        values = [data[key].get(m, 0) for m in months]

        # Create bar chart
        bars = ax.bar(month_labels, values, color=color, alpha=0.85, edgecolor='white', linewidth=1.5)

        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'¥{height:,.0f}',
                       ha='center', va='bottom', fontsize=10, fontweight='bold')

        # Formatting - All Chinese
        ax.set_title(title, fontsize=14, fontweight='bold', pad=12)
        ax.set_xlabel('月份', fontsize=11)
        ax.set_ylabel('金额 (元)', fontsize=11)
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.tick_params(axis='x', rotation=0)

        # Format y-axis
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{int(x):,}'))

    plt.tight_layout()

    # Save to file
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'月度分类进展_{timestamp}.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"✓ 月度分类进展图已生成: {output_file}")
    return output_file


if __name__ == '__main__':
    # Test
    generate_monthly_category_chart()
