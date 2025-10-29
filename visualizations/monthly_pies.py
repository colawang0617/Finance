"""
Monthly Percentage Distribution - Pie Charts Grid

Generates a grid of pie charts showing percentage breakdown of venue income
subcategories for each month
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib
from datetime import datetime
import os

# Set Chinese font support - Use Heiti TC for bold, rounded style
matplotlib.rcParams['font.sans-serif'] = ['Heiti TC', 'STHeiti', 'Microsoft YaHei', 'SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False
matplotlib.rcParams['font.size'] = 12
matplotlib.rcParams['font.weight'] = 'medium'  # Bolder default weight


def get_monthly_venue_breakdown(filepath, months):
    """
    Get venue income breakdown by subcategory for each month

    Args:
        filepath: Path to Excel file
        months: List of months to include

    Returns:
        Dictionary with monthly breakdown data
    """
    # Use data_only=True to get calculated values instead of formulas
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['每日数据']

    # Initialize data structure for venue subcategories
    # Columns: C=Meituan, D=Card Deduct, E=Douyin, F=Coach Fee, G=WeChat, H=Alipay
    monthly_data = {}

    for month in months:
        monthly_data[month] = {
            'meituan': 0,
            'card_deduct': 0,
            'douyin': 0,
            'coach_fee': 0,
            'wechat': 0,
            'alipay': 0
        }

    # Collect data
    for row in range(3, ws.max_row + 1):
        date_val = ws[f'A{row}'].value
        if not date_val:
            continue

        try:
            # Handle both string dates ("08-18") and datetime objects
            if isinstance(date_val, str):
                month = int(date_val.split('-')[0])
            else:
                # It's a datetime object
                month = date_val.month

            if month not in months:
                continue

            # Get subcategory values
            meituan = ws[f'C{row}'].value or 0
            card_deduct = ws[f'D{row}'].value or 0
            douyin = ws[f'E{row}'].value or 0
            coach_fee = ws[f'F{row}'].value or 0
            wechat = ws[f'G{row}'].value or 0
            alipay = ws[f'H{row}'].value or 0

            # Accumulate
            monthly_data[month]['meituan'] += meituan if isinstance(meituan, (int, float)) else 0
            monthly_data[month]['card_deduct'] += card_deduct if isinstance(card_deduct, (int, float)) else 0
            monthly_data[month]['douyin'] += douyin if isinstance(douyin, (int, float)) else 0
            monthly_data[month]['coach_fee'] += coach_fee if isinstance(coach_fee, (int, float)) else 0
            monthly_data[month]['wechat'] += wechat if isinstance(wechat, (int, float)) else 0
            monthly_data[month]['alipay'] += alipay if isinstance(alipay, (int, float)) else 0

        except (ValueError, IndexError):
            continue

    wb.close()
    return monthly_data


def generate_monthly_pies(months=[5, 6, 7, 8, 9, 10], filepath='Finance/财务跟踪表_完整版_KL.xlsx', output_dir='reports/graphs'):
    """
    Generate grid of monthly pie charts showing venue income distribution

    Args:
        months: List of months to display
        filepath: Path to Excel file
        output_dir: Directory to save output PNG

    Returns:
        Path to generated PNG file
    """
    # Get data
    data = get_monthly_venue_breakdown(filepath, months)

    # Filter out months with no data
    data = {m: v for m, v in data.items() if sum(v.values()) > 0}

    if not data:
        print("✗ 没有找到数据")
        return None

    # Subcategory labels and colors - Chinese only
    labels = ['美团', '储值卡核销', '抖音', '教练课核销', '微信', '支付宝']

    colors = ['#5DADE2', '#C39BD3', '#F4A460', '#CD6155', '#82B366', '#64B5F6']

    month_names = {
        5: '5月', 6: '6月', 7: '7月',
        8: '8月', 9: '9月', 10: '10月'
    }

    # Create figure with subplots (2 rows x 3 columns)
    fig, axes = plt.subplots(2, 3, figsize=(18, 12))
    fig.suptitle('场地入账 - 月度百分比分布 (2025)',
                 fontsize=22, fontweight='bold', y=0.995)

    # Plot each month
    for idx, month in enumerate(sorted(data.keys())):
        row = idx // 3
        col = idx % 3
        ax = axes[row, col]

        month_data = data[month]
        values = [
            month_data['meituan'],
            month_data['card_deduct'],
            month_data['douyin'],
            month_data['coach_fee'],
            month_data['wechat'],
            month_data['alipay']
        ]

        total = sum(values)

        # Only show categories with non-zero values
        filtered_labels = []
        filtered_values = []
        filtered_colors = []

        for label, value, color in zip(labels, values, colors):
            if value > 0:
                filtered_labels.append(label)
                filtered_values.append(value)
                filtered_colors.append(color)

        # Create pie chart
        wedges, texts, autotexts = ax.pie(
            filtered_values,
            labels=filtered_labels,
            colors=filtered_colors,
            autopct='%1.1f%%',
            startangle=90,
            textprops={'fontsize': 10, 'fontweight': 'bold'},
            wedgeprops={'edgecolor': 'white', 'linewidth': 2}
        )

        # Make percentage text white
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontsize(11)
            autotext.set_fontweight('bold')

        # Title with month name and total - Chinese only
        ax.set_title(f'{month_names.get(month, f"{month}月")} 2025\n总计: ¥{total:,.0f}',
                    fontsize=14, fontweight='bold', pad=15)

    # Remove empty subplots if we have less than 6 months
    for idx in range(len(data), 6):
        row = idx // 3
        col = idx % 3
        fig.delaxes(axes[row, col])

    # Add legend at the bottom
    fig.legend(labels, loc='lower center', ncol=6, fontsize=11,
              bbox_to_anchor=(0.5, -0.02), frameon=True, fancybox=True)

    plt.tight_layout(rect=[0, 0.03, 1, 0.98])

    # Save to file
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'月度百分比分布_{timestamp}.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"✓ 月度百分比分布图已生成: {output_file}")
    return output_file


if __name__ == '__main__':
    # Test
    generate_monthly_pies()
