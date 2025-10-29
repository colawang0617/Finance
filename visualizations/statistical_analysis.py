"""
Statistical Analysis Chart - Venue Income

Generates comprehensive statistical analysis showing key metrics
for venue income across all months.
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.patches import Rectangle
from datetime import datetime
import os
import numpy as np

# Set Chinese font support - Use Heiti TC for bold, rounded style
matplotlib.rcParams['font.sans-serif'] = ['Heiti TC', 'STHeiti', 'Microsoft YaHei', 'SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False
matplotlib.rcParams['font.size'] = 12
matplotlib.rcParams['font.weight'] = 'medium'


def get_monthly_venue_data(filepath, months):
    """
    Get venue income data for each month

    Args:
        filepath: Path to Excel file
        months: List of months to analyze

    Returns:
        Dictionary with monthly data arrays
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['每日数据']

    monthly_data = {month: [] for month in months}

    # Collect data
    for row in range(3, ws.max_row + 1):
        date_val = ws[f'A{row}'].value
        if not date_val:
            continue

        try:
            # Handle both string dates and datetime objects
            if isinstance(date_val, str):
                month = int(date_val.split('-')[0])
            else:
                month = date_val.month

            if month not in months:
                continue

            venue = ws[f'B{row}'].value
            if venue and isinstance(venue, (int, float)) and venue > 0:
                monthly_data[month].append(venue)

        except (ValueError, IndexError, AttributeError):
            continue

    wb.close()
    return monthly_data


def calculate_statistics(data_array):
    """
    Calculate comprehensive statistics for a data array

    Args:
        data_array: List of numeric values

    Returns:
        Dictionary with statistical metrics
    """
    if not data_array or len(data_array) == 0:
        return None

    arr = np.array(data_array)

    return {
        'count': len(arr),
        'mean': np.mean(arr),
        'median': np.median(arr),
        'std': np.std(arr, ddof=1),  # Sample std dev
        'min': np.min(arr),
        'max': np.max(arr),
        'q25': np.percentile(arr, 25),
        'q75': np.percentile(arr, 75),
        'range': np.max(arr) - np.min(arr),
        'cv': (np.std(arr, ddof=1) / np.mean(arr) * 100) if np.mean(arr) > 0 else 0
    }


def generate_statistical_analysis(months=[5,6,7,8,9,10], filepath='Finance/财务跟踪表_完整版_KL.xlsx', output_dir='reports/graphs'):
    """
    Generate comprehensive statistical analysis chart for venue income

    Args:
        months: List of months to analyze
        filepath: Path to Excel file
        output_dir: Directory to save output PNG

    Returns:
        Path to generated PNG file
    """
    # Get data
    monthly_data = get_monthly_venue_data(filepath, months)

    # Filter out empty months
    monthly_data = {m: v for m, v in monthly_data.items() if len(v) > 0}

    if not monthly_data:
        print("✗ 没有找到数据")
        return None

    # Calculate statistics for each month
    monthly_stats = {}
    for month, data in monthly_data.items():
        monthly_stats[month] = calculate_statistics(data)

    # Create figure with 3 panels
    fig = plt.figure(figsize=(20, 14))
    gs = fig.add_gridspec(3, 2, height_ratios=[1.2, 1.1, 1.2], hspace=0.35, wspace=0.3)

    fig.suptitle('场地入账 - 月度统计分析 (2025)', fontsize=24, fontweight='bold', y=0.98)

    # Month names
    month_names = {5: '5月', 6: '6月', 7: '7月', 8: '8月', 9: '9月', 10: '10月'}

    # ========================
    # Panel 1: Box Plot (Top, spans 2 columns)
    # ========================
    ax1 = fig.add_subplot(gs[0, :])

    positions = list(range(len(monthly_data)))
    box_data = [monthly_data[m] for m in sorted(monthly_data.keys())]
    labels = [month_names.get(m, f'{m}月') for m in sorted(monthly_data.keys())]

    bp = ax1.boxplot(box_data, positions=positions, widths=0.6,
                     patch_artist=True,
                     boxprops=dict(facecolor='#5DADE2', alpha=0.7, linewidth=2),
                     medianprops=dict(color='#FF6B6B', linewidth=3),
                     whiskerprops=dict(linewidth=1.5),
                     capprops=dict(linewidth=1.5),
                     flierprops=dict(marker='o', markerfacecolor='red', markersize=8, alpha=0.5))

    ax1.set_xticks(positions)
    ax1.set_xticklabels(labels, fontsize=14, fontweight='bold')
    ax1.set_ylabel('金额 (元)', fontsize=14, fontweight='bold')
    ax1.set_title('月度箱线图 - 显示中位数、四分位数和异常值', fontsize=16, fontweight='bold', pad=15)
    ax1.grid(axis='y', alpha=0.3, linestyle='--')
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'¥{x:.2f}'))

    # ========================
    # Panel 2: Statistics Table (Middle left)
    # ========================
    ax2 = fig.add_subplot(gs[1, 0])
    ax2.axis('tight')
    ax2.axis('off')

    # Build table data
    table_data = []
    table_data.append(['月份', '样本数', '平均值', '中位数', '标准差', '最小值', '最大值'])

    for month in sorted(monthly_data.keys()):
        stats = monthly_stats[month]
        table_data.append([
            month_names.get(month, f'{month}月'),
            f"{stats['count']}天",
            f"¥{stats['mean']:.2f}",
            f"¥{stats['median']:.2f}",
            f"¥{stats['std']:.2f}",
            f"¥{stats['min']:.2f}",
            f"¥{stats['max']:.2f}"
        ])

    table = ax2.table(cellText=table_data, cellLoc='center', loc='center',
                     colWidths=[0.11, 0.11, 0.17, 0.17, 0.16, 0.15, 0.15])

    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 2.8)

    # Style header row
    for i in range(7):
        cell = table[(0, i)]
        cell.set_facecolor('#5DADE2')
        cell.set_text_props(weight='bold', color='white', fontsize=12)

    # Alternate row colors
    for i in range(1, len(table_data)):
        for j in range(7):
            cell = table[(i, j)]
            if i % 2 == 0:
                cell.set_facecolor('#E8F4F8')
            else:
                cell.set_facecolor('white')

    ax2.set_title('基础统计指标', fontsize=14, fontweight='bold', pad=20, y=1.05)

    # ========================
    # Panel 3: Advanced Statistics Table (Middle right)
    # ========================
    ax3 = fig.add_subplot(gs[1, 1])
    ax3.axis('tight')
    ax3.axis('off')

    # Build advanced stats table
    adv_table_data = []
    adv_table_data.append(['月份', '25分位', '75分位', '极差', '变异系数'])

    for month in sorted(monthly_data.keys()):
        stats = monthly_stats[month]
        adv_table_data.append([
            month_names.get(month, f'{month}月'),
            f"¥{stats['q25']:.2f}",
            f"¥{stats['q75']:.2f}",
            f"¥{stats['range']:.2f}",
            f"{stats['cv']:.2f}%"
        ])

    adv_table = ax3.table(cellText=adv_table_data, cellLoc='center', loc='center',
                         colWidths=[0.14, 0.21, 0.21, 0.22, 0.22])

    adv_table.auto_set_font_size(False)
    adv_table.set_fontsize(10)
    adv_table.scale(1, 2.8)

    # Style header row
    for i in range(5):
        cell = adv_table[(0, i)]
        cell.set_facecolor('#F4A460')
        cell.set_text_props(weight='bold', color='white', fontsize=12)

    # Alternate row colors
    for i in range(1, len(adv_table_data)):
        for j in range(5):
            cell = adv_table[(i, j)]
            if i % 2 == 0:
                cell.set_facecolor('#FFF4E6')
            else:
                cell.set_facecolor('white')

    ax3.set_title('高级统计指标', fontsize=14, fontweight='bold', pad=20, y=1.05)

    # ========================
    # Panel 4: Monthly Comparison Bar Chart (Bottom, spans 2 columns)
    # ========================
    ax4 = fig.add_subplot(gs[2, :])

    sorted_months = sorted(monthly_data.keys())
    x = np.arange(len(sorted_months))
    width = 0.25

    means = [monthly_stats[m]['mean'] for m in sorted_months]
    medians = [monthly_stats[m]['median'] for m in sorted_months]
    maxs = [monthly_stats[m]['max'] for m in sorted_months]

    bars1 = ax4.bar(x - width, means, width, label='平均值', color='#5DADE2', alpha=0.85, edgecolor='white', linewidth=2)
    bars2 = ax4.bar(x, medians, width, label='中位数', color='#F4A460', alpha=0.85, edgecolor='white', linewidth=2)
    bars3 = ax4.bar(x + width, maxs, width, label='最大值', color='#CD6155', alpha=0.85, edgecolor='white', linewidth=2)

    # Add value labels
    for bars in [bars1, bars2, bars3]:
        for bar in bars:
            height = bar.get_height()
            ax4.text(bar.get_x() + bar.get_width()/2., height,
                    f'¥{height:.2f}',
                    ha='center', va='bottom', fontsize=9, fontweight='bold')

    ax4.set_xlabel('月份', fontsize=14, fontweight='bold')
    ax4.set_ylabel('金额 (元)', fontsize=14, fontweight='bold')
    ax4.set_title('月度统计对比 - 平均值、中位数、最大值', fontsize=16, fontweight='bold', pad=15)
    ax4.set_xticks(x)
    ax4.set_xticklabels([month_names.get(m, f'{m}月') for m in sorted_months], fontsize=13)
    ax4.legend(fontsize=12, loc='upper left', framealpha=0.9)
    ax4.grid(axis='y', alpha=0.3, linestyle='--')
    ax4.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'¥{x:.2f}'))

    plt.tight_layout()

    # Save to file
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'统计分析_场地_{timestamp}.png')
    plt.savefig(output_file, dpi=300, bbox_inches='tight')
    plt.close()

    print(f"✓ 统计分析图已生成: {output_file}")
    return output_file


if __name__ == '__main__':
    # Test
    generate_statistical_analysis()
